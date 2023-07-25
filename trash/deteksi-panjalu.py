from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.select import Select
import hashlib
import pyautogui
import time as t
from openpyxl import load_workbook

wb = load_workbook(filename='dt-panjalu.xlsx')
sheet = wb['Sheet1']

options = Options() 

options.add_argument('--user-data-dir=C:/Users/Irpan/AppData/Local/Google/Chrome/User Data')
options.add_argument('--profile-directory=Default')
options.add_experimental_option('detach', True)

driver = webdriver.Chrome(options=options)

driver.get("https://bos.polri.go.id/laporan/deteksi-dini")
wait = WebDriverWait(driver, timeout=3000)

def wait_until(xpath):
    wait.until(expected_conditions.visibility_of_element_located((By.XPATH, xpath)))

def select_el(x_path):
    return driver.find_element(By.XPATH, x_path)

def set_select(xpath, data_value):
    selectEl = select_el(xpath)
    select = Select(selectEl)
    select.select_by_value(data_value)

try: 
    wait2 = WebDriverWait(driver, timeout=2)
    wait2.until(expected_conditions.visibility_of_element_located((By.XPATH, '/html/body/main/div[2]/div/a')))
except:
    print('data sudaah ada')


wb = load_workbook(filename='list.xlsx')
sheet = wb['Sheet1']
i=2
while i <= 10 :
# loop
# variable definition
    print('setting variable')
    nama = sheet['C' + str(i)].value
    pekerjaan = "wiraswasta"
    tanggal = sheet['B' + str(i)].value
    jam = '10:00'
    dusun = sheet['D' + str(i)].value
    rt = sheet['E' + str(i)].value
    rw = sheet['F' + str(i)].value
    desa = sheet['G' + str(i)].value
    keterangan = sheet['H' + str(i)].value
    keterangan = keterangan.replace('$nama', nama) 
    keterangan = keterangan.replace('$dusun', dusun) 
    keterangan = keterangan.replace('$rt', rt) 
    keterangan = keterangan.replace('$rw', rw) 
    keterangan = keterangan.replace('$desa', desa) 

    i+=1

    print('menunggu klik tombol tambah')
    wait_until('//*[@id="table-deteksi-dini_length"]/a')
    select_el('//*[@id="table-deteksi-dini_length"]/a').click()

    # expand all form
    try:
        select_el('/html/body/main/div/form/span[3]').click()
        select_el('/html/body/main/div/form/span[2]').click()
        select_el('/html/body/main/div/form/span[1]').click()
    except:
        print('gagal expand element')

    # fill form 1
    try: 
        wait_until('//*[@id="nama_narasumber"]')
        select_el('//*[@id="nama_narasumber"]').send_keys(nama)
        select_el('//*[@id="pekerjaan"]').send_keys(pekerjaan)

        # select alamat
        # select provinsi
        set_select('//*[@id="provinsi"]', 'JAWA BARAT')
        # select kabupaten 
        wait_until('//*[@id="3207"]')
        set_select('//*[@id="kabupaten"]', 'KABUPATEN CIAMIS')
        # select KECAMATAN 
        wait_until('//*[@id="320708 "]')
        set_select('//*[@id="kecamatan"]', 'PANJALU')
        # select desa
        wait_until('//*[@id="3207082004"]')
        set_select('//*[@id="desa"]', desa)

        # isi dusun rt rw
        select_el('//*[@id="detail_alamat"]').send_keys(dusun)
        select_el('//*[@id="rt"]').send_keys(rt)
        select_el('//*[@id="rw"]').send_keys(rw)
    except:
        print('gagal fil form 1')


    # isi form 2
    try: 
        select_el('//*[@id="tanggal"]').send_keys(tanggal)
        select_el('//*[@id="jam_mendapatkan_informasi"]').send_keys(jam)
        select_el('//*[@id="lokasi_mendapatkan_informasi"]').send_keys('RUMAH NARASUMBER')
    except:
        print('gagal isi form 2')

    # isi form 3
    try:
        driver.execute_script('document.getElementsByName("laporan_informasi[bidang]")[1].click()')
        select_el('//*[@id="uraian_informasi"]').send_keys(keterangan)
        ActionChains(driver).send_keys(Keys.TAB*2).send_keys('ekonomi').send_keys(Keys.ENTER).perform()
    except:
        print('gagal klik radio')


    # submit
    
    driver.execute_script('document.getElementsByTagName("form")[2].submit()')
    wait_until('/html/body/div[3]/div/div[3]/button[1]')
    select_el('/html/body/div[3]/div/div[3]/button[1]').click()

    # end loop 
